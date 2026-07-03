from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from database import get_db
from models import Student, User, UserRole, Submission, PublishabilityStatus, Comment
from schemas import StudentListItem, SupervisorDashboardStats, BulkImportResponse, StudentImportRow, SubmissionOut, CommentOut, CommentCreate, PublishabilityUpdate, SupervisorNotesUpdate
from dependencies import require_supervisor, get_current_user
from auth import hash_password
import openpyxl
from typing import List

router = APIRouter(prefix="/supervisor", tags=["Supervisor"], dependencies=[Depends(require_supervisor)])

@router.get("/students", response_model=list[StudentListItem])
def list_students(db: Session = Depends(get_db)):
    students = db.query(Student).all()
    return students

@router.get("/dashboard", response_model=SupervisorDashboardStats)
def dashboard_stats(db: Session = Depends(get_db)):
    total_students = db.query(Student).count()
    total_submissions = db.query(Submission).count()
    pending = db.query(Submission).filter(Submission.id.notin_(
        db.query(Submission.id).join(Comment, Comment.submission_id == Submission.id)
    )).count()
    publishable = db.query(Student).filter(Student.publishability_status == PublishabilityStatus.PUBLISHABLE).count()
    
    return SupervisorDashboardStats(
        total_students=total_students,
        total_submissions=total_submissions,
        pending_reviews=pending,
        publishable_count=publishable,
    )

@router.post("/upload-students", response_model=BulkImportResponse)
async def upload_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files allowed")
    
    contents = await file.read()
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active
    
    headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
    
    # Map columns
    col_map = {}
    for idx, h in enumerate(headers):
        if h in ["name", "full name", "student name"]:
            col_map["name"] = idx
        elif h in ["matric", "matric number", "matriculation number", "reg no"]:
            col_map["matric"] = idx
        elif h in ["project title", "title", "project"]:
            col_map["title"] = idx
        elif h in ["department", "dept"]:
            col_map["dept"] = idx
    
    imported = 0
    failed: List[StudentImportRow] = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            name = str(row[col_map.get("name", 0)] or "").strip()
            matric = str(row[col_map.get("matric", 1)] or "").strip()
            title = str(row[col_map.get("title", 2)] or "").strip()
            dept = str(row[col_map.get("dept", 3)] or "").strip()
            
            if not all([name, matric, title, dept]):
                failed.append(StudentImportRow(
                    name=name, matricNumber=matric, projectTitle=title, department=dept,
                    isValid=False, errors=["Missing required fields"]
                ))
                continue
            
            # Check duplicate
            if db.query(Student).filter(Student.matric_number == matric).first():
                failed.append(StudentImportRow(
                    name=name, matricNumber=matric, projectTitle=title, department=dept,
                    isValid=False, errors=["Matric number already exists"]
                ))
                continue
            
            # Create user + student
            user = User(
                name=name,
                identifier=matric,
                hashed_password=hash_password("Caleb123"),  # default
                role=UserRole.STUDENT,
                is_first_login=True,
            )
            db.add(user)
            db.flush()  # get user.id
            
            student = Student(
                user_id=user.id,
                name=name,
                matric_number=matric,
                project_title=title,
                department=dept,
            )
            db.add(student)
            imported += 1
            
        except Exception as e:
            failed.append(StudentImportRow(
                name=str(row[0] or ""), matricNumber=str(row[1] or ""),
                projectTitle=str(row[2] or ""), department=str(row[3] or ""),
                isValid=False, errors=[str(e)]
            ))
    
    db.commit()
    return BulkImportResponse(imported=imported, failed=failed)

@router.get("/student/{student_id}", response_model=StudentListItem)
def get_student_detail(student_id: str, db: Session = Depends(get_db)):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student

@router.get("/student/{student_id}/submissions", response_model=list[SubmissionOut])
def get_student_submissions(student_id: str, db: Session = Depends(get_db)):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    submissions = db.query(Submission).filter(Submission.student_id == student_id).all()
    for s in submissions:
        s.commentCount = db.query(Comment).filter(Comment.submission_id == s.id).count()
    return submissions

@router.put("/student/{student_id}/publishability", response_model=StudentListItem)
def update_publishability(
    student_id: str,
    req: PublishabilityUpdate,
    db: Session = Depends(get_db),
):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    student.publishability_status = req.status
    db.commit()
    db.refresh(student)
    return student

@router.put("/student/{student_id}/notes", response_model=StudentListItem)
def update_supervisor_notes(
    student_id: str,
    req: SupervisorNotesUpdate,
    db: Session = Depends(get_db),
):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    student.supervisor_notes = req.notes
    db.commit()
    db.refresh(student)
    return student

@router.get("/submissions/{submission_id}/comments", response_model=list[CommentOut])
def get_submission_comments(submission_id: str, db: Session = Depends(get_db)):
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    comments = db.query(Comment).filter(Comment.submission_id == submission_id).order_by(Comment.created_at.asc()).all()
    return comments

@router.post("/submissions/{submission_id}/comments", response_model=CommentOut)
def add_submission_comment(
    submission_id: str,
    req: CommentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    comment = Comment(
        submission_id=submission_id,
        author_name=req.author_name or current_user.name,
        body=req.body,
    )
    db.add(comment)
    db.commit()
    db.refresh(comment)
    return comment
